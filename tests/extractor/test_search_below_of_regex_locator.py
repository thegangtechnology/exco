import pytest
from openpyxl import Workbook

from exco import CellLocation
from exco.cell_full_path import CellFullPath
from exco.extractor.locator.built_in.search_below_of_regex_locator import SearchBelowOfRegexLocator
from exco.extractor.locator.locating_result import LocatingResult


@pytest.fixture
def wb() -> CellFullPath:
    wb = Workbook()

    ws = wb.active
    ws['A6'] = 40
    ws['A3'] = 'the key'

    return wb


def test_search_below_of_regex_locator_pass(wb: Workbook):
    rol = SearchBelowOfRegexLocator(regex='^.*key$', max_empty_row_search=5)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A3"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="A6"
    )
    assert result == LocatingResult.good(cell_loc)


def test_search_below_of_regex_locator_max(wb: Workbook):
    rol = SearchBelowOfRegexLocator(regex='^.*key$', max_empty_row_search=2)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A3"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="A5"
    )
    assert result == LocatingResult.good(cell_loc)


def test_search_right_of_regex_locator_fail(wb: Workbook):
    rol = SearchBelowOfRegexLocator(regex='^.*k$', max_empty_row_search=2)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A3"
    ), workbook=wb)
    assert result == LocatingResult.bad(
            msg="Unable to find cell below of ^.*k$")
