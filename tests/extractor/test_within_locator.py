import pytest
from openpyxl.workbook import Workbook

from exco import CellLocation
from exco.cell_full_path import CellFullPath
from exco.extractor.locator.built_in.within_locator import WithinLocator
from exco.extractor.locator.locating_result import LocatingResult

# All of this to fix code smells
test_table_name = 'test table'
test_table_2_name = 'test table 2'
test_table_3_name = 'test table 3'

repeated_cell_name = 'same name here'
unable_to_find = 'unable to find me'
below_of_name = 'below of'
single_cell_name = 'single cell'
non_existent_cell = 'non existent'
right_of_name = 'right of'
cell_above_everything = 'top hidden cell'

# directions
right_of_direction = "right_of"
below_of_direction = 'below_of'


def _create_error_message(table_name: str = None,
                          cell_name: str = None,
                          direction: str = None,
                          label: str = None) -> str:
    if label is not None:
        return "Unable to find cell with label " + label
    return "Unable to find cell " + cell_name + " to the " + direction.replace('_', ' ') + " " + table_name


@pytest.fixture
def wb() -> CellFullPath:
    wb = Workbook()

    ws = wb.active

    # Single Cell Checks
    ws['K2'] = single_cell_name
    ws['K10'] = below_of_name
    ws['K11'] = 10
    ws['L10'] = 1
    ws['P2'] = right_of_name
    ws['Q2'] = 14
    ws['P3'] = 15
    ws['L2'] = single_cell_name
    ws['M2'] = 1
    ws['K3'] = single_cell_name
    ws['K4'] = 7
    ws['L3'] = 1

    # Merged Cells Check
    ws['D2'] = cell_above_everything

    ws.merge_cells('C3:E19')
    ws['C3'] = test_table_name

    ws['G6'] = repeated_cell_name
    ws['H6'] = 1
    ws['G7'] = 2

    ws['F3'] = 'top part'
    ws['G3'] = 2

    ws['F19'] = 'below part'
    ws['G19'] = 456

    ws['F13'] = test_table_name
    ws['G13'] = 16

    ws['F20'] = unable_to_find  # is exactly 1 cell between merged cells "test table" and "test table 2"

    ws.merge_cells('C21:E32')
    ws['C21'] = test_table_2_name

    ws['G22'] = repeated_cell_name
    ws['H22'] = 2
    ws['G23'] = 1

    ws['D36'] = below_of_name
    ws['D37'] = 8
    ws['E36'] = 9

    ws['C33'] = 'left part'
    ws['C34'] = 9

    ws['E33'] = 'right part'
    ws['E34'] = 10

    ws['D33'] = test_table_2_name
    ws['D34'] = 11

    ws.merge_cells('C51:E57')
    ws['C51'] = test_table_3_name

    ws['D60'] = below_of_name
    ws['D61'] = 12
    ws['E60'] = 15

    return wb


def test_within_locator_no_such_label_on_table_fail(wb: Workbook):
    rol = WithinLocator(direction=right_of_direction, label='fake table', find=repeated_cell_name,
                        perform=right_of_direction)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A3"
    ), workbook=wb)
    assert result == LocatingResult.bad(msg=_create_error_message(label='fake table'))


def test_within_locator_no_such_perform_on_table_fail(wb: Workbook):
    rol = WithinLocator(direction=right_of_direction, label=test_table_name, find=repeated_cell_name, perform='top_of')
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A3"
    ), workbook=wb)
    assert result == LocatingResult.bad(msg="Incorrect perform, must be one of the following ['"
                                            + right_of_direction + "', '" + below_of_direction + "']")


def test_within_locator_no_such_direction_fail(wb: Workbook):
    rol = WithinLocator(direction='top_of', label=test_table_name, find=repeated_cell_name, perform=right_of_direction)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A3"
    ), workbook=wb)
    assert result == LocatingResult.bad(
        msg="Incorrect direction, must be one of the following ['"
            + right_of_direction + "', '" + below_of_direction + "']")


# RIGHT OF
def test_within_right_of(wb: Workbook):
    rol = WithinLocator(direction=right_of_direction, label=test_table_name, find=repeated_cell_name,
                        perform=right_of_direction)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A3"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="H6"
    )
    assert result == LocatingResult.good(cell_loc)

    rol = WithinLocator(direction=right_of_direction, label=test_table_name, find=repeated_cell_name,
                        perform=below_of_direction)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A3"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="G7"
    )
    assert result == LocatingResult.good(cell_loc)


def test_within_right_of_single_cell(wb: Workbook):
    rol = WithinLocator(direction=right_of_direction, label=single_cell_name, find=right_of_name,
                        perform=right_of_direction)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="Z1"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="Q2"
    )
    assert result == LocatingResult.good(cell_loc)

    rol = WithinLocator(direction=right_of_direction, label=single_cell_name, find=right_of_name,
                        perform=below_of_direction)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="Z1"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="P3"
    )
    assert result == LocatingResult.good(cell_loc)


def test_within_right_of_single_cell_fail(wb: Workbook):
    rol = WithinLocator(direction=right_of_direction, label=single_cell_name, find=non_existent_cell,
                        perform=right_of_direction)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="Z1"
    ), workbook=wb)
    assert result == LocatingResult.bad(msg=_create_error_message(table_name=single_cell_name,
                                                                  cell_name=non_existent_cell,
                                                                  direction=right_of_direction))


def test_within_right_of_boundary_single_cell(wb: Workbook):
    rol = WithinLocator(direction=right_of_direction, label=single_cell_name, find=single_cell_name,
                        perform=right_of_direction)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="Z1"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="L2"
    )
    assert result == LocatingResult.good(cell_loc)

    rol = WithinLocator(direction=right_of_direction, label=single_cell_name, find=single_cell_name,
                        perform=below_of_direction)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="Z1"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="K3"
    )
    assert result == LocatingResult.good(cell_loc)


def test_within_right_of_fail(wb: Workbook):
    rol = WithinLocator(direction=right_of_direction, label=test_table_name, find=right_of_name,
                        perform=right_of_direction)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A10"
    ), workbook=wb)
    assert result == LocatingResult.bad(msg=_create_error_message(table_name=test_table_name,
                                                                  cell_name=right_of_name,
                                                                  direction=right_of_direction))

    rol = WithinLocator(direction=right_of_direction, label=test_table_name, find=right_of_name,
                        perform=below_of_direction)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A10"
    ), workbook=wb)
    assert result == LocatingResult.bad(msg=_create_error_message(table_name=test_table_name,
                                                                  cell_name=right_of_name,
                                                                  direction=right_of_direction))


def test_within_right_of_boundary_top_check_fail(wb: Workbook):
    rol = WithinLocator(direction=right_of_direction, label=test_table_2_name, find=unable_to_find,
                        perform=right_of_direction)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A5"
    ), workbook=wb)
    assert result == LocatingResult.bad(msg=_create_error_message(table_name=test_table_2_name,
                                                                  cell_name=unable_to_find,
                                                                  direction=right_of_direction))

    rol = WithinLocator(direction=right_of_direction, label=test_table_2_name, find=unable_to_find,
                        perform=below_of_direction)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A5"
    ), workbook=wb)
    assert result == LocatingResult.bad(msg=_create_error_message(table_name=test_table_2_name,
                                                                  cell_name=unable_to_find,
                                                                  direction=right_of_direction))


def test_within_right_of_boundary_below_check_fail(wb: Workbook):
    rol = WithinLocator(direction=right_of_direction, label=test_table_name, find=unable_to_find,
                        perform=right_of_direction)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A4"
    ), workbook=wb)
    assert result == LocatingResult.bad(msg=_create_error_message(table_name=test_table_name,
                                                                  cell_name=unable_to_find,
                                                                  direction=right_of_direction))

    rol = WithinLocator(direction=right_of_direction, label=test_table_name, find=unable_to_find,
                        perform=below_of_direction)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A4"
    ), workbook=wb)
    assert result == LocatingResult.bad(msg=_create_error_message(table_name=test_table_name,
                                                                  cell_name=unable_to_find,
                                                                  direction=right_of_direction))


def test_within_same_find_right_of(wb: Workbook):
    rol = WithinLocator(direction=right_of_direction, label=test_table_name, find=repeated_cell_name,
                        perform=right_of_direction)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A6"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="H6"
    )
    assert result == LocatingResult.good(cell_loc)

    rol = WithinLocator(direction=right_of_direction, label=test_table_2_name, find=repeated_cell_name,
                        perform=right_of_direction)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A7"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="H22"
    )
    assert result == LocatingResult.good(cell_loc)

    rol = WithinLocator(direction=right_of_direction, label=test_table_name, find=repeated_cell_name,
                        perform=below_of_direction)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A3"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="G7"
    )
    assert result == LocatingResult.good(cell_loc)

    rol = WithinLocator(direction=right_of_direction, label=test_table_2_name, find=repeated_cell_name,
                        perform=below_of_direction)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A3"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="G23"
    )
    assert result == LocatingResult.good(cell_loc)


# below OF
def test_within_below_of_single_cell_fail(wb: Workbook):
    rol = WithinLocator(direction=below_of_direction, label=single_cell_name, find=non_existent_cell,
                        perform="below_of")
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="Z1"
    ), workbook=wb)
    assert result == LocatingResult.bad(msg=_create_error_message(table_name=single_cell_name,
                                                                  cell_name=non_existent_cell,
                                                                  direction=below_of_direction))

    rol = WithinLocator(direction=below_of_direction, label=single_cell_name, find=non_existent_cell,
                        perform="right_of")
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="Z1"
    ), workbook=wb)
    assert result == LocatingResult.bad(msg=_create_error_message(table_name=single_cell_name,
                                                                  cell_name=non_existent_cell,
                                                                  direction=below_of_direction))


def test_within_below_of_single_cell(wb: Workbook):
    rol = WithinLocator(direction=below_of_direction, label=single_cell_name, find=below_of_name, perform="below_of")
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="Z1"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="K11"
    )
    assert result == LocatingResult.good(cell_loc)

    rol = WithinLocator(direction=below_of_direction, label=single_cell_name, find=below_of_name, perform="right_of")
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="Z1"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="L10"
    )
    assert result == LocatingResult.good(cell_loc)


def test_within_below_of_boundary_single_cell(wb: Workbook):
    rol = WithinLocator(direction=below_of_direction, label=single_cell_name, find=single_cell_name, perform="below_of")
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="Z1"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="K4"
    )
    assert result == LocatingResult.good(cell_loc)

    rol = WithinLocator(direction=below_of_direction, label=single_cell_name, find=single_cell_name, perform="right_of")
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="Z1"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="L3"
    )
    assert result == LocatingResult.good(cell_loc)


def test_within_below_of(wb: Workbook):
    rol = WithinLocator(direction=below_of_direction, label=test_table_name, find=below_of_name, perform="below_of")
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="B3"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="D37"
    )
    assert result == LocatingResult.good(cell_loc)

    rol = WithinLocator(direction=below_of_direction, label=test_table_name, find=below_of_name, perform="right_of")
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="B3"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="E36"
    )
    assert result == LocatingResult.good(cell_loc)


def test_within_below_of_fail(wb: Workbook):
    rol = WithinLocator(direction=below_of_direction, label=test_table_name, find=right_of_name, perform="below_of")
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="B10"
    ), workbook=wb)
    assert result == LocatingResult.bad(msg=_create_error_message(table_name=test_table_name,
                                                                  cell_name=right_of_name,
                                                                  direction=below_of_direction))

    rol = WithinLocator(direction=below_of_direction, label=test_table_name, find=right_of_name, perform="right_of")
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="B10"
    ), workbook=wb)
    assert result == LocatingResult.bad(msg=_create_error_message(table_name=test_table_name,
                                                                  cell_name=right_of_name,
                                                                  direction=below_of_direction))


def test_within_below_of_boundary_top_check_fail(wb: Workbook):
    rol = WithinLocator(direction=below_of_direction, label=test_table_2_name, find=unable_to_find, perform="below_of")
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="B5"
    ), workbook=wb)
    assert result == LocatingResult.bad(msg=_create_error_message(table_name=test_table_2_name,
                                                                  cell_name=unable_to_find,
                                                                  direction=below_of_direction))

    rol = WithinLocator(direction=below_of_direction, label=test_table_name,
                        find=cell_above_everything,
                        perform="below_of")
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="B5"
    ), workbook=wb)
    assert result == LocatingResult.bad(
        msg=_create_error_message(table_name=test_table_name, cell_name=cell_above_everything,
                                  direction=below_of_direction))

    rol = WithinLocator(direction=below_of_direction, label=test_table_2_name, find=unable_to_find, perform="right_of")
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="B5"
    ), workbook=wb)
    assert result == LocatingResult.bad(msg=_create_error_message(table_name=test_table_2_name,
                                                                  cell_name=unable_to_find,
                                                                  direction=below_of_direction))

    rol = WithinLocator(direction=below_of_direction, label=test_table_name,
                        find=cell_above_everything,
                        perform="right_of")
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="B5"
    ), workbook=wb)
    assert result == LocatingResult.bad(
        msg=_create_error_message(table_name=test_table_name, cell_name=cell_above_everything,
                                  direction=below_of_direction))


def test_within_below_of_boundary_below_check_fail(wb: Workbook):
    rol = WithinLocator(direction=below_of_direction, label=test_table_name, find=unable_to_find, perform="below_of")
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="B4"
    ), workbook=wb)
    assert result == LocatingResult.bad(msg=_create_error_message(table_name=test_table_name,
                                                                  cell_name=unable_to_find,
                                                                  direction=below_of_direction))

    rol = WithinLocator(direction=below_of_direction, label=test_table_name, find=unable_to_find, perform="right_of")
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="B4"
    ), workbook=wb)
    assert result == LocatingResult.bad(msg=_create_error_message(table_name=test_table_name,
                                                                  cell_name=unable_to_find,
                                                                  direction=below_of_direction))


def test_within_same_find_below_of(wb: Workbook):
    rol = WithinLocator(direction=below_of_direction, label=test_table_name, find=below_of_name, perform="below_of")
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="B6"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="D37"
    )
    assert result == LocatingResult.good(cell_loc)

    rol = WithinLocator(direction=below_of_direction, label=test_table_3_name, find=below_of_name, perform="below_of")
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="B7"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="D61"
    )
    assert result == LocatingResult.good(cell_loc)

    rol = WithinLocator(direction=below_of_direction, label=test_table_name, find=below_of_name, perform="right_of")
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="B6"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="E36"
    )
    assert result == LocatingResult.good(cell_loc)

    rol = WithinLocator(direction=below_of_direction, label=test_table_3_name, find=below_of_name, perform="right_of")
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="B7"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="E60"
    )
    assert result == LocatingResult.good(cell_loc)
