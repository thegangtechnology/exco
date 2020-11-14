from unittest.mock import patch

import pytest
from openpyxl import Workbook
from openpyxl.cell import Cell

from exco.cell_full_path import CellFullPath
from exco.extractor import Assumption
from exco.extractor.assumption.assumption_result import AssumptionResult
from exco.extractor.assumption.built_in.left_cell_match_assumption import LeftCellMatchAssumption


@pytest.fixture
def cfp() -> CellFullPath:
    wb = Workbook()

    ws = wb.active
    ws['B2'] = 30
    ws['A2'] = 'the key'

    cell = Cell(row=2, column=2, worksheet=wb)
    cell_full_path = CellFullPath(
        workbook=wb, sheet=wb.active, cell=cell)
    return cell_full_path


@patch.multiple(Assumption, __abstractmethods__=set())
def test_assumption_abstract(cfp):
    with pytest.raises(NotImplementedError):
        assumption = Assumption()
        assumption.assume(cfp=cfp)


def test_assumption_result(cfp):
    lcm_assumption = LeftCellMatchAssumption(label='the key')
    assert lcm_assumption.assume(cfp=cfp) == AssumptionResult.good()

    lcm_assumption = LeftCellMatchAssumption(label='the key 2')
    assert lcm_assumption.assume(cfp=cfp) == AssumptionResult.bad(
        'Cell to the left of B2 does not match the key 2')
