import pytest
from openpyxl import Workbook

from exco import CellLocation
from exco.cell_full_path import CellFullPath
from exco.extractor.locator.built_in.search_right_of_locator import SearchRightOfLocator
from exco.extractor.locator.locating_result import LocatingResult

A3_CELL_VALUE = 'the key 1'


@pytest.fixture
def wb() -> CellFullPath:
    wb = Workbook()

    ws = wb.active
    ws['D3'] = 40
    ws['A3'] = A3_CELL_VALUE

    return wb


def test_search_right_of_locator(wb: Workbook):
    rol = SearchRightOfLocator(label=A3_CELL_VALUE, max_empty_col_search=5)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A3"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="D3"
    )
    assert result == LocatingResult.good(cell_loc)


def test_search_right_of_locator_max(wb: Workbook):
    rol = SearchRightOfLocator(label=A3_CELL_VALUE, max_empty_col_search=2)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A3"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="C3"
    )
    assert result == LocatingResult.good(cell_loc)


def test_search_right_of_locator_fail(wb: Workbook):
    rol = SearchRightOfLocator(label='the key', max_empty_col_search=2)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A3"
    ), workbook=wb)
    assert result == LocatingResult.bad(
        msg='Unable to find cell to the right of the key')
