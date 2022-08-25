import pytest
from openpyxl import Workbook

from exco import CellLocation
from exco.cell_full_path import CellFullPath
from exco.extractor.locator.built_in.below_of_locator import BelowOfLocator
from exco.extractor.locator.locating_result import LocatingResult


@pytest.fixture
def wb() -> CellFullPath:
    wb = Workbook()

    ws = wb.active
    ws['A6'] = 40
    ws['A3'] = 'the key 1'

    ws.merge_cells('E4:H4')
    ws['E5'] = 20
    ws['E4'] = 'horizontal merged cell'

    ws.merge_cells('E6:H11')
    ws['E12'] = 321
    ws['E6'] = 'boxed merged cell'

    ws.merge_cells('B6:B11')
    ws['B12'] = 123
    ws['B6'] = 'vertical merged cell'

    return wb


def test_below_of_locator_fail(wb: Workbook):
    rol = BelowOfLocator(label='the key')
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A3"
    ), workbook=wb)
    assert result == LocatingResult.bad(
        msg='Unable to find cell below of the key')


def test_below_of_locator_merged_cell_fail(wb: Workbook):
    rol = BelowOfLocator(label='vertical merged')
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="F3"
    ), workbook=wb)
    assert result == LocatingResult.bad(
        msg='Unable to find cell below of vertical merged')


def test_below_of_locator(wb: Workbook):
    rol = BelowOfLocator(label='the key 1')
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="B3"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="A4"
    )
    assert result == LocatingResult.good(cell_loc)


def test_below_of_locator_horizontal_merged_cell(wb: Workbook):
    rol = BelowOfLocator(label='horizontal merged cell')
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A5"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="E5"
    )
    assert result == LocatingResult.good(cell_loc)


def test_below_of_locator_vertical_merged_cell(wb: Workbook):
    rol = BelowOfLocator(label='vertical merged cell')
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A6"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="B12"
    )
    assert result == LocatingResult.good(cell_loc)


def test_below_of_locator_boxed_merged_cell(wb: Workbook):
    rol = BelowOfLocator(label='boxed merged cell')
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A7"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="E12"
    )
    assert result == LocatingResult.good(cell_loc)

