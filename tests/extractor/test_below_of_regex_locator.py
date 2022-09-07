import pytest
from openpyxl import Workbook

from exco import CellLocation
from exco.cell_full_path import CellFullPath
from exco.extractor.locator.built_in.below_of_regex_locator import BelowOfRegexLocator
from exco.extractor.locator.locating_result import LocatingResult


@pytest.fixture
def wb() -> CellFullPath:
    wb = Workbook()

    ws = wb.active
    ws['A6'] = 40
    ws['A3'] = 'the key 1'

    ws['B9'] = 50
    ws['B7'] = 're-geXx'

    return wb


def test_below_of_regex_locator_fail(wb: Workbook):
    rol = BelowOfRegexLocator(regex='^T.*1$')
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A3"
    ), workbook=wb)
    assert result == LocatingResult.bad(
        msg='Unable to find cell below of ^T.*1$')


def test_below_of_regex_locator_pass(wb: Workbook):
    rol = BelowOfRegexLocator(regex='^r.*x$', n=2)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="B5"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="B9"
    )
    assert result == LocatingResult.good(cell_loc)
