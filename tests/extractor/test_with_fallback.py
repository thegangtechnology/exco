from openpyxl import Workbook
from openpyxl.comments import Comment

from exco import util, ExcoTemplate, ExcelProcessorFactory


def test_with_defaults():
    wb = Workbook()
    sheet = wb.active
    sheet.cell(1, 1).value = 'not a number'
    sheet.cell(1, 1).comment = Comment(util.long_string("""
        {{--
        key: key
        parser: int
        fallback: 999
        --}}
    """), 'author')

    spec = ExcoTemplate.from_workbook(wb).to_raw_excel_processor_spec()
    print(spec)
    ep = ExcelProcessorFactory.default().create_from_spec(spec)

    result = ep.process_workbook(wb)
    assert not result.is_ok
    assert result.to_dict() == {'key': 999}
