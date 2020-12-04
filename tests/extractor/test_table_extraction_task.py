import openpyxl
import pytest
from exco import CellLocation
from exco.cell_location import CellOffset
from exco.exception import NoEndConditionError, TooManyRowRead
from exco.extractor.cell_extraction_task import CellExtractionTask
from exco.extractor.locator.built_in.at_comment_cell_locator import AtCommentCellLocator
from exco.extractor.locator.built_in.right_of_locator import RightOfLocator
from exco.extractor.parser.built_in.string_parser import StringParser
from exco.extractor.table_end_conditions.built_in.all_blank_table_end_condition import AllBlankTableEndCondition
from exco.extractor.table_end_conditions.built_in.max_row_table_end_condition import MaxRowTableEndCondition
from exco.extractor.table_end_conditions.table_end_condition_factory import TableEndConditionFactory
from exco.extractor.table_extraction_task import EndConditionCollection, TableExtractionTask
from exco.extractor_spec.table_extraction_spec import TableItemDirection
from openpyxl import Workbook


def test_end_condition_collection():
    with pytest.raises(NoEndConditionError):
        EndConditionCollection.from_spec(
            [],
            TableEndConditionFactory.default()
        )


@pytest.fixture(scope='function')
def workbook() -> Workbook:
    wb = openpyxl.Workbook()
    sheet = wb.active
    for i in range(10):
        sheet.cell(i + 1, 1).value = i
    return wb


@pytest.fixture(scope='function')
def rightward_table_wb() -> Workbook:
    wb = openpyxl.Workbook()
    sheet = wb.active
    for i in range(10):
        sheet.cell(1, i + 1).value = i
    return wb


@pytest.fixture(scope='function')
def cell_loc() -> CellLocation:
    return CellLocation(sheet_name='Sheet', coordinate='A1')


def test_extract_rightward_table(rightward_table_wb, cell_loc):
    tt = TableExtractionTask(
        key="some_table",
        locator=AtCommentCellLocator(),
        columns={CellOffset(row=0, col=0): CellExtractionTask.simple(
            key='some_key', parser=StringParser())},
        end_condition=EndConditionCollection([AllBlankTableEndCondition()]),
        item_direction=TableItemDirection.RIGHTWARD
    )
    result = tt.process(cell_loc, rightward_table_wb)
    assert len(result.row_results) == 10


def test_table_extraction_task(workbook, cell_loc):
    tt = TableExtractionTask(
        key="some_table",
        locator=AtCommentCellLocator(),
        columns={CellOffset(row=0, col=0): CellExtractionTask.simple(
            key='some_key', parser=StringParser())},
        end_condition=EndConditionCollection([MaxRowTableEndCondition(n=3)]),
        item_direction=TableItemDirection.DOWNWARD
    )

    result = tt.process(cell_loc, workbook)
    assert len(result.row_results) == 3


def test_table_extraction_task_hit_infinite_loop_guard(workbook, cell_loc):
    tt = TableExtractionTask(
        key="some_table",
        locator=AtCommentCellLocator(),
        columns={CellOffset(row=0, col=0): CellExtractionTask.simple(
            key='some_key', parser=StringParser())},
        end_condition=EndConditionCollection([]),
        item_direction=TableItemDirection.DOWNWARD
    )

    with pytest.raises(TooManyRowRead):
        tt.process(cell_loc, workbook)


def test_table_extraction_fail_locating(workbook, cell_loc):
    tt = TableExtractionTask(
        key="some_table",
        locator=RightOfLocator(label='Non Existent'),
        columns={CellOffset(0, 0): CellExtractionTask.simple(
            key='some_key', parser=StringParser())},
        end_condition=EndConditionCollection.default(),
        item_direction=TableItemDirection.DOWNWARD
    )

    result = tt.process(cell_loc, workbook)
    assert not result.locating_result.is_ok


def test_shift_cell_downward(cell_loc):
    tt = TableExtractionTask(
        key="some_table",
        locator=AtCommentCellLocator(),
        columns={CellOffset(0, 0): CellExtractionTask.simple(
            key='some_key', parser=StringParser())},
        end_condition=EndConditionCollection.default(),
        item_direction=TableItemDirection.DOWNWARD
    )
    cl = CellLocation(sheet_name='Sheet', coordinate='A1')
    assert tt.shift_column_direction(cl, 1).coordinate == 'B1'
    assert tt.shift_item_direction(cl, 1).coordinate == 'A2'


def test_shift_cell_rightward(cell_loc):
    tt = TableExtractionTask(
        key="some_table",
        locator=AtCommentCellLocator(),
        columns={CellOffset(0, 0): CellExtractionTask.simple(
            key='some_key', parser=StringParser())},
        end_condition=EndConditionCollection.default(),
        item_direction=TableItemDirection.RIGHTWARD
    )
    cl = CellLocation(sheet_name='Sheet', coordinate='A1')
    assert tt.shift_column_direction(cl, 1).coordinate == 'A2'
    assert tt.shift_item_direction(cl, 1).coordinate == 'B1'
