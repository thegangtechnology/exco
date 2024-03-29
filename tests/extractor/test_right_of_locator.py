import pytest
from openpyxl import Workbook

from exco import CellLocation
from exco.cell_full_path import CellFullPath
from exco.extractor.locator.built_in.right_of_locator import RightOfLocator
from exco.extractor.locator.built_in.right_of_regex_locator import RightOfRegexLocator
from exco.extractor.locator.locating_result import LocatingResult

E4_CELL_VALUE = 'right of horizontal merged cell'
E6_CELL_VALUE = 'right of boxed merged cell'
B6_CELL_VALUE = 'right of vertical merged cell'


@pytest.fixture
def wb() -> CellFullPath:
    wb = Workbook()

    ws = wb.active
    ws['B3'] = 40
    ws['A3'] = 'right of'

    ws['C4'] = 50
    ws['B4'] = 'the key 1'

    ws.merge_cells('E4:H4')
    ws['I4'] = 20
    ws['E4'] = E4_CELL_VALUE

    ws.merge_cells('E6:H11')
    ws['I6'] = 321
    ws['E6'] = E6_CELL_VALUE

    ws.merge_cells('B6:B11')
    ws['C6'] = 123
    ws['B6'] = B6_CELL_VALUE

    return wb


def test_right_of_locator_fail(wb: Workbook):
    rol = RightOfLocator(label='the key')
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A3"
    ), workbook=wb)
    assert result == LocatingResult.bad(
        msg='Unable to find cell to the right of the key')


def test_right_of_locator_merged_cell_fail(wb: Workbook):
    rol = RightOfLocator(label='right of vertical merged')
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="F3"
    ), workbook=wb)
    assert result == LocatingResult.bad(
        msg='Unable to find cell to the right of right of vertical merged')


def test_right_of_locator(wb: Workbook):
    rol = RightOfLocator(label='right of')
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A4"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="B3"
    )
    assert result == LocatingResult.good(cell_loc)


def test_right_of_locator_horizontal_merged_cell(wb: Workbook):
    rol = RightOfLocator(label=E4_CELL_VALUE)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A5"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="I4"
    )
    assert result == LocatingResult.good(cell_loc)


def test_right_of_locator_vertical_merged_cell(wb: Workbook):
    rol = RightOfLocator(label=B6_CELL_VALUE)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A6"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="C6"
    )
    assert result == LocatingResult.good(cell_loc)


def test_right_of_locator_boxed_merged_cell(wb: Workbook):
    rol = RightOfLocator(label=E6_CELL_VALUE)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A7"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="I6"
    )
    assert result == LocatingResult.good(cell_loc)


def test_right_of_locator_regex(wb: Workbook):
    rol = RightOfRegexLocator(regex='the key\\s\\d')
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A8"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="C4"
    )
    assert result == LocatingResult.good(cell_loc)


def test_right_of_locator_regex_failed(wb: Workbook):
    rol = RightOfRegexLocator(regex='the key\\s\\ds')
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A9"
    ), workbook=wb)
    assert result == LocatingResult.bad(
        msg='Unable to find cell to the right of the key\\s\\ds')


def test_right_of_locator_regex_boxed_merged_cell(wb: Workbook):
    rol = RightOfRegexLocator(regex=E6_CELL_VALUE)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A10"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="I6"
    )
    assert result == LocatingResult.good(cell_loc)


def test_right_of_locator_regex_vertical_merged_cell(wb: Workbook):
    rol = RightOfRegexLocator(regex=B6_CELL_VALUE)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A10"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="C6"
    )
    assert result == LocatingResult.good(cell_loc)


def test_right_of_locator_regex_horizontal_merged_cell(wb: Workbook):
    rol = RightOfRegexLocator(regex=E4_CELL_VALUE)
    result = rol.locate(anchor_cell_location=CellLocation(
        sheet_name="Sheet",
        coordinate="A5"
    ), workbook=wb)
    cell_loc = CellLocation(
        sheet_name="Sheet",
        coordinate="I4"
    )
    assert result == LocatingResult.good(cell_loc)
