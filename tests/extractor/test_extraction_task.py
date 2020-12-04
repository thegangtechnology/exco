from os.path import join, dirname

import exco
import pytest
from openpyxl import Workbook

from exco import CellLocation
from exco.extractor.assumption.assumption_result import AssumptionResult
from exco.extractor.assumption.built_in.left_cell_match_assumption import LeftCellMatchAssumption
from exco.extractor.cell_extraction_task import CellExtractionTaskResult, CellExtractionTask
from exco.extractor.locator.built_in.at_comment_cell_locator import AtCommentCellLocator
from exco.extractor.locator.built_in.right_of_locator import RightOfLocator
from exco.extractor.locator.locating_result import LocatingResult
from exco.extractor.parser.built_in.int_parser import IntParser


@pytest.fixture
def wb() -> object:
    wb = Workbook()

    ws = wb.active
    ws['A2'] = 'the key'
    ws['B2'] = 30
    return wb


def test_extraction_task_result_failed():
    ctr = CellExtractionTaskResult.fail_locating(key="something",
                                                 locating_result=LocatingResult(location=None,
                                                                                is_ok=False),
                                                 fallback='funny',
                                                 metadata={})
    assert not ctr.is_ok
    assert ctr.get_value() == 'funny'


def test_extraction_task_assumption_failed():
    ctr = CellExtractionTaskResult.fail_assumptions(key="something",
                                                    locating_result=LocatingResult(
                                                        location=None, is_ok=False),
                                                    assumption_results={
                                                        "a": AssumptionResult.bad(msg="failed")
                                                    },
                                                    fallback='fallback',
                                                    metadata={})
    assert not ctr.is_ok
    assert ctr.get_value() == 'fallback'


def test_extraction_task_locating_failed(wb: Workbook):
    et = CellExtractionTask(
        key="h1",
        locator=RightOfLocator(label="hi"),
        parser=IntParser(),
        validators={},
        assumptions={},
        fallback='fallback',
        metadata={}
    )

    result = et.process(anchor_cell_location=CellLocation(sheet_name="Sheet",
                                                          coordinate="A1"),
                        workbook=wb)
    assert not result.is_ok
    assert result.get_value() == 'fallback'


def test_extraction_task_assumption(wb: Workbook):
    et = CellExtractionTask(
        key="h1",
        locator=AtCommentCellLocator(),
        parser=IntParser(),
        validators={},
        assumptions={"B2": LeftCellMatchAssumption(label="the koi")},
        fallback='fallback',
        metadata={}
    )

    result = et.process(anchor_cell_location=CellLocation(sheet_name="Sheet",
                                                          coordinate="B2"),
                        workbook=wb)
    assert not result.is_ok
    assert result.get_value() == 'fallback'


def test_extraction_with_metadata():
    fname = join(dirname(__file__), '../../sample/test/simple_with_meta.xlsx')
    processor = exco.from_excel(fname)

    result = processor.process_excel(fname)
    assert result.cell_result_for_key('distance').result.metadata['unit'] == 'km'
