from dataclasses import dataclass
from typing import Dict, List

from excel_comment_orm.cell_location import CellLocation
from excel_comment_orm.eco_template.eco_block import ECOBlock
import openpyxl as opx
from excel_comment_orm.exception import ECOException, BadTemplateException, CommentWithNoECOBlockWarning
from excel_comment_orm.extraction_spec.excel_processor_spec import ExcelProcessorSpec


@dataclass
class ECOTemplate:
    eco_blocks: Dict[CellLocation, List[ECOBlock]]

    def n_cell(self) -> int:
        """

        Returns:
            Total number of cells with comments. Some comment may have no eco block.
        """
        return len(self.eco_blocks)

    def n_eco_blocks(self) -> int:
        """

        Returns:
            Total number of ECOBlocks
        """
        return sum(len(block) for block in self.eco_blocks.values())

    def cells_with_no_eco_block(self) -> List[CellLocation]:
        """

        Returns:
            List[CellLocation] where it has comments but no ECO Block.
        """
        return [k for k, v in self.eco_blocks.items() if len(v) == 0]

    @classmethod
    def from_workbook(cls, workbook: opx.Workbook) -> 'ECOTemplate':
        from openpyxl.worksheet.worksheet import Worksheet
        ret = {}
        wb = workbook
        for sheet_name in wb.sheetnames:
            sheet: Worksheet = wb[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.comment is not None:
                        cell_loc = CellLocation(sheet_name=sheet_name, coordinate=cell.coordinate)
                        try:
                            ret[cell_loc] = ECOBlock.from_string(cell.comment.text)
                        except ECOException as e:  # throw error with cell info
                            raise BadTemplateException(  # Todo: maybe be all these should be warning
                                f'Bad Template at sheet: {cell_loc.sheet_name}, cell: {cell_loc.coordinate}') from e

        et = ECOTemplate(ret)
        if et.cells_with_no_eco_block():
            questionable_cells = et.cells_with_no_eco_block()
            raise CommentWithNoECOBlockWarning("Found Cell with comment but no eco block.\n"
                                               f"[{', '.join(x.short_name for x in questionable_cells)}]")
        return et

    @classmethod
    def from_excel(cls, fname: str) -> 'ECOTemplate':
        return cls.from_workbook(workbook=opx.load_workbook(fname))

    def to_excel_extractor_spec(self) -> ExcelProcessorSpec:
        ret = {}
        for cell_loc, eco_blocks in self.eco_blocks.items():
            spec = []
            for block in eco_blocks:
                try:
                    spec.append(block.to_extractor_task_spec())
                except ECOException as e:
                    raise BadTemplateException(f'Fail to construct spec at {cell_loc.short_name}\n'
                                               f'{block.raw}') from e
            ret[cell_loc] = spec

        return ExcelProcessorSpec(ret)