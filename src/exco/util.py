import itertools
import textwrap
from collections import defaultdict
from datetime import date
from typing import TypeVar, Iterable, Any, Dict, List, Tuple, Type, Set, Optional, Generator, Callable, Union

import openpyxl
import stringcase
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.merge import MergedCellRange
from openpyxl.worksheet.worksheet import Worksheet

from exco import setting as st
from exco.cell_full_path import CellFullPath

T = TypeVar('T')
CellValue = Union[str, int, date, None]
CellLocation = 'CellLocation'  # to avoid flake8 code smell


def long_string(s: str) -> str:
    """left strip and dedent the string

    Args:
        s (str):

    Returns:
        left string and dedent the string
    """
    return textwrap.dedent(s).lstrip()


def is_unique(items: Iterable[Any]) -> bool:
    s = set()
    for item in items:
        if item in s:
            return False
        s.add(item)
    return True


def tuple_to_coordinate(row, col) -> str:
    return get_column_letter(col) + str(row)


def coordinate_to_tuple(coord: str) -> Tuple[int, int]:
    return openpyxl.utils.coordinate_to_tuple(coord)


def shift_coord(coord: str, shift: Tuple[int, int]):
    row, col = coordinate_to_tuple(coord)
    sr, sc = shift
    return tuple_to_coordinate(row + sr, col + sc)


def remove_suffix(s: str, suffix: str):  # polyfill for python<3.9
    """Remove suffix(if exists) from s

    Args:
        s (str):
        suffix (str): suffix to remove.

    Returns:
        suffix removed s.
    """
    if s.endswith(suffix):
        return s[:-len(suffix)]
    else:
        return s


def default_key(clz: Type[Any], suffix):
    """Return snake case with suffix removed. This is used as default key for class.

    Args:
        clz (Type[Any]): class
        suffix (str): suffix to remove

    Returns:
        default key name: Ex: IntParser -> int

    """
    return stringcase.snakecase(remove_suffix(clz.__name__, suffix))


def extra_keys(d: Dict[str, Any], allowed=Set[str]) -> List[str]:
    return [k for k in d.keys() if k not in allowed]


def name_params(d: Dict[str, Any],
                exclude: Optional[Set[str]] = None) -> Tuple[str, Dict[str, Any]]:
    exclude = set() if exclude is None else exclude
    name = d[st.k_name]
    params = {k: v for k, v in d.items() if k != st.k_name and k not in exclude}
    return name, params


def flatten(lol: Iterable[Iterable[T]]) -> Iterable[T]:
    return itertools.chain.from_iterable(lol)


def flattened_len(it: Iterable[Iterable]):
    """Find the len of flatten list
    """
    return sum(1 for _ in flatten(it))


def iterate_cells_in_worksheet(
        sheet: Worksheet) -> Generator[Cell, None, None]:
    for row in sheet.iter_rows():
        for cell in row:
            yield cell


def iterate_cells_in_workbook(
        workbook: Workbook) -> Generator[CellFullPath, None, None]:
    for sheetname in workbook.sheetnames:
        sheet: Worksheet = workbook[sheetname]
        for cell in iterate_cells_in_worksheet(sheet):
            yield CellFullPath(
                workbook=workbook,
                sheet=sheet,
                cell=cell)


def unique(xs: Iterable[T]) -> List[T]:
    return list(set(xs))


T2 = TypeVar('T2')


def group_by(f: Callable[[T], T2], xs: List[T]) -> Dict[T2, List[T]]:
    ret = defaultdict(list)
    for x in xs:
        key = f(x)
        ret[key].append(x)
    return dict(ret)


def get_merged_cell(sheet: Worksheet, coordinates: str) -> Optional[MergedCellRange]:
    """Find the merge cell in the whole sheet that contains coordinates

    Args:
        sheet (Worksheet): worksheet
        coordinates (str): coordinates of cell

    Returns:
        MergedCellRange if cell is a part of a merged cell, None if cell is not a merged cell
    """
    for merged_cell in sheet.merged_cells.ranges:
        if coordinates in merged_cell:
            return merged_cell
    return None


def get_rightmost_coordinate(sheet: Worksheet, cell: Cell) -> CellLocation:
    from exco import CellLocation
    merged_cell = get_merged_cell(sheet=sheet, coordinates=cell.coordinate)
    if merged_cell is None:
        return CellLocation(
            coordinate=cell.coordinate,
            sheet_name=sheet.title
        )
    return CellLocation(
        coordinate=tuple_to_coordinate(cell.row, merged_cell.max_col),
        sheet_name=sheet.title
    )


def get_bottommost_coordinate(sheet: Worksheet, cell: Cell) -> CellLocation:
    from exco import CellLocation
    merged_cell = get_merged_cell(sheet=sheet, coordinates=cell.coordinate)
    if merged_cell is None:
        return CellLocation(
            coordinate=cell.coordinate,
            sheet_name=sheet.title
        )
    return CellLocation(
        coordinate=tuple_to_coordinate(merged_cell.max_row, cell.column),
        sheet_name=sheet.title
    )


def iter_rows_between(sheet: Worksheet, cell: Cell) -> Generator[Cell, None, None]:
    """Loops over the cells to the right of the rows between cell (inclusive). This works for cells and merged cells.

    Args:
        sheet (Worksheet): worksheet
        cell (Cell): Cell

    Returns:
        Generator of Cells
    """
    merged_cell = get_merged_cell(sheet=sheet, coordinates=cell.coordinate)
    cell_range = coordinate_to_tuple(cell.coordinate) if merged_cell is None else merged_cell
    if type(cell_range) is MergedCellRange:
        min_col = cell_range.max_col
        min_row = cell_range.min_row
        max_col = sheet.max_column
        max_row = cell_range.max_row
    else:
        min_col = cell_range[1]
        min_row = cell_range[0]
        max_col = sheet.max_column
        max_row = cell_range[0]
    for row in range(min_row, max_row + 1):
        cells = (sheet.cell(row=row, column=column) for column in range(min_col, max_col + 1))
        yield tuple(cells)


def iter_cols_between(sheet: Worksheet, cell: Cell) -> Generator[Cell, None, None]:
    """Loops over the cells beneath the columns between cell (inclusive). This works for cells and merged cells.

    Args:
        sheet (Worksheet): worksheet
        cell (Cell): Cell

    Returns:
        Generator of Cells
    """
    merged_cell = get_merged_cell(sheet=sheet, coordinates=cell.coordinate)
    cell_range = coordinate_to_tuple(cell.coordinate) if merged_cell is None else merged_cell
    if type(cell_range) is MergedCellRange:
        min_col = cell_range.min_col
        min_row = cell_range.max_row
        max_col = cell_range.max_col
        max_row = sheet.max_row
    else:
        min_col = cell_range[1]
        min_row = cell_range[0]
        max_col = cell_range[1]
        max_row = sheet.max_row
    for row in range(min_row + 1, max_row + 1):
        cells = (sheet.cell(row=row, column=column) for column in range(min_col, max_col + 1))
        yield tuple(cells)


def search_right_of_scope(sheet: Worksheet, cell: Cell, label: str) -> Optional[Cell]:
    """Loop over the cells to the right of the rows between cell (inclusive) and
    find the cell with the specified label.

    Args:
        sheet (Worksheet): worksheet
        cell (Cell): Cell
        label (str): cell name to find

    Returns:
        Generator of Cells
    """
    for row in iter_rows_between(sheet=sheet, cell=cell):
        for cell in row:
            if cell.value == label:
                return cell
    return None


def search_below_of_scope(sheet: Worksheet, cell: Cell, label: str) -> Optional[Cell]:
    """Loop over the cells beneath the columns between cell (inclusive) and
    find the cell with the specified label.

    Args:
        sheet (Worksheet): worksheet
        cell (Cell): Cell
        label (str): cell name to find

    Returns:
        Generator of Cells
    """
    for col in iter_cols_between(sheet=sheet, cell=cell):
        for cell in col:
            if cell.value == label:
                return cell
    return None
