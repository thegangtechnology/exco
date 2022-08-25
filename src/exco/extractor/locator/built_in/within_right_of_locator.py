from dataclasses import dataclass
from typing import Optional, Union, Generator

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.merge import MergedCellRange
from openpyxl.worksheet.worksheet import Worksheet

from exco import CellLocation, util
from exco.extractor.locator.locating_result import LocatingResult
from exco.extractor.locator.locator import Locator

CellRange = Union[MergedCellRange, CellLocation]

@dataclass
class WithinRightOfLocator(Locator):
    label: str
    find: str

    def locate(self, anchor_cell_location: CellLocation,
               workbook: Workbook) -> LocatingResult:
        sheet: Worksheet = workbook[anchor_cell_location.sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == self.label:
                    found_cell = self._search_right_of_cell(sheet=sheet, cell=cell)
                    if found_cell is None:
                        return LocatingResult.bad(
                            msg=f"Unable to find cell {self.find} to the right of {self.label}")
                    coord = util.get_rightmost_coordinate(sheet=sheet, cell=found_cell)
                    cell_loc = CellLocation(
                        sheet_name=anchor_cell_location.sheet_name,
                        coordinate=util.shift_coord(coord.coordinate, (0, 1))
                    )
                    return LocatingResult.good(cell_loc)
        return LocatingResult.bad(
            msg=f"Unable to find cell with label {self.label}")

    def _search_right_of_cell(self, sheet: Worksheet, cell: Cell) -> Optional[Cell]:
        merged_cell = util.get_merged_cell(sheet=sheet, coordinates=cell.coordinate)
        merged_cell = util.coordinate_to_tuple(cell.coordinate) if merged_cell is None else merged_cell
        return self._right_of_scope(sheet=sheet, cell_range=merged_cell)

    def _right_of_scope(self, sheet: Worksheet, cell_range: CellRange) -> Optional[Cell]:
        # search only rows of merged cell's or cell's column range
        for row in self._cells_by_row_between(sheet=sheet, cell_range=cell_range):
            for cell in row:
                cell_coordinates = util.coordinate_to_tuple(cell.coordinate)
                if type(cell_range) is MergedCellRange:
                    # to maintain that it is to the right of merged cell and within the column range
                    if cell_range.min_row <= cell_coordinates[0] <= cell_range.max_row and \
                            cell_coordinates[1] >= cell_range.max_col:
                        if cell.value == self.find:
                            return cell
                else:
                    if cell_range.col == cell_coordinates[1]:
                        if cell.value == self.find:
                            return cell
        return None

    def _cells_by_row_between(self, sheet: Worksheet, cell_range: CellRange) -> Generator[Cell, None, None]:
        if type(cell_range) is MergedCellRange:
            min_col = cell_range.max_col
            min_row = cell_range.min_row
            max_col = sheet.max_column
            max_row = cell_range.max_row
        else:
            min_col = cell_range.col
            min_row = cell_range.row
            max_col = cell_range.col
            max_row = cell_range.row
        for row in range(min_row, max_row + 1):
            cells = (sheet.cell(row=row, column=column) for column in range(min_col, max_col + 1))
            yield tuple(cells)
