import secrets
from dataclasses import dataclass
from typing import TypeVar, Dict, Any, List, Optional, Generic, Type

import openpyxl
from openpyxl import Workbook

from exco.cell_location import CellLocation
from exco.exception import ExcoException, ExtractionTaskCreationException, TableExtractionTaskCreationException
from exco.extractor.assumption.assumption import Assumption
from exco.extractor.assumption.assumption_factory import AssumptionFactory
from exco.extractor.cell_extraction_task import CellExtractionTaskResult, CellExtractionTask
from exco.extractor.locator.locator import Locator
from exco.extractor.locator.locator_factory import LocatorFactory
from exco.extractor.parser.parser import Parser
from exco.extractor.parser.parser_factory import ParserFactory
from exco.extractor.table_end_conditions.table_end_condition import TableEndCondition
from exco.extractor.table_end_conditions.table_end_condition_factory import TableEndConditionFactory
from exco.extractor.table_extraction_task import TableExtractionTask, EndConditionCollection, TableExtractionTaskResult
from exco.extractor.validator.validator import Validator
from exco.extractor.validator.validator_factory import ValidatorFactory
from exco.extractor_spec import CellExtractionSpec, ExcelProcessorSpec
from exco.extractor_spec.table_extraction_spec import TableExtractionSpec
from exco.sheet_name_alias import SheetName, SheetNameAliasCheckers

T = TypeVar('T')


@dataclass
class ProcessorKey:
    cell_location: CellLocation
    key: str

    def __hash__(self):
        return hash((self.cell_location, self.key))


@dataclass
class LookupResult(Generic[T]):
    cell_location: CellLocation
    result: T


@dataclass
class ExcelProcessingResult:
    cell_results: Dict[CellLocation, List[CellExtractionTaskResult]]
    table_results: Dict[CellLocation, List[TableExtractionTaskResult]]

    @property
    def is_ok(self):
        return all(cr.is_ok for crs in self.cell_results.values() for cr in crs) and \
               all(tr.is_ok for trs in self.table_results.values() for tr in trs)

    @staticmethod
    def _lookup_for_key(d: Dict[CellLocation,
                                List[T]],
                        key: str) -> Optional[LookupResult[T]]:
        for cl, results in d.items():
            for result in results:
                if key == result.key:
                    return LookupResult(cl, result)
            # TODO: Fix this O(n) search.
        return None

    def cell_result_for_key(
            self, key: str) -> Optional[LookupResult[CellExtractionTaskResult]]:
        return self._lookup_for_key(self.cell_results, key)

    def table_result_for_key(
            self, key: str) -> Optional[LookupResult[TableExtractionTaskResult]]:
        return self._lookup_for_key(self.table_results, key)

    def to_dict(self) -> Dict[str, Any]:
        ret = {}
        for results in self.cell_results.values(
        ):  # TODO: Warn of duplicate key
            for result in results:
                ret[result.key] = result.get_value()

        for results in self.table_results.values(
        ):  # TODO: Warn of duplicate key
            for result in results:
                ret[result.key] = result.get_value()
        return ret


@dataclass
class ExcelDerefedProcessor:
    cell_processors: Dict[CellLocation, List[CellExtractionTask]]
    table_processors: Dict[CellLocation, List[TableExtractionTask]]

    def process_workbook(self, workbook: Workbook) -> ExcelProcessingResult:
        cell_result = {}
        for loc, cets in self.cell_processors.items():
            cell_result[loc] = [cet.process(loc, workbook) for cet in cets]

        table_result = {}
        for loc, tets in self.table_processors.items():
            table_result[loc] = [tet.process(loc, workbook) for tet in tets]
        return ExcelProcessingResult(
            cell_results=cell_result, table_results=table_result)

    def process_excel(self, fname: str) -> ExcelProcessingResult:
        wb = openpyxl.load_workbook(fname, data_only=True)
        return self.process_workbook(wb)

    def __str__(self):
        tmp = []
        for cl, tasks in self.cell_processors.items():
            header = f'Location: {cl.short_name}\n'
            sep = '*\n' * 10
            task_str = sep.join([str(task) + '\n' for task in tasks])
            tmp.append(header + task_str)
        ret = ("#" * 20 + '\n').join(tmp)
        return ret


@dataclass
class ExcelProcessor:
    spec: ExcelProcessorSpec  # raw spec
    factory: 'ExcelProcessorFactory'
    sheet_name_checkers: SheetNameAliasCheckers
    accept_only_visible_sheets: bool

    def deref(self, workbook: Optional[Workbook]) -> ExcelDerefedProcessor:
        if workbook is not None:
            derefed_spec = self.spec.spec_to_extractor_deref(workbook)
            return self.factory.create_derefed_processor_from_spec(derefed_spec)
        else:
            return self.factory.create_derefed_processor_from_spec(self.spec)

    def process_workbook(self, workbook: Workbook) -> ExcelProcessingResult:
        workbook = self.normalize_workbook_sheet_names(workbook)
        return self.deref(workbook).process_workbook(workbook)

    def deal_with_duplicates(self, workbook: Workbook, template_sheet_names: List[str]) -> Workbook:
        hidden_sheet_names = [sheet.title for sheet in workbook.worksheets if sheet.sheet_state == "hidden"]
        for template_sheet_name in template_sheet_names:
            if template_sheet_name in hidden_sheet_names:
                workbook[template_sheet_name].title = "duplicate_sheet_name" + str(secrets.randbelow(10000))
        return workbook

    def normalize_workbook_sheet_names(self, workbook: Workbook) -> Workbook:
        if not self.sheet_name_checkers:
            return workbook
        if self.accept_only_visible_sheets:
            workbook = self.deal_with_duplicates(workbook, list(self.sheet_name_checkers.keys()))
        name_mapping: Dict[SheetName, SheetName] = {}
        for sheet in workbook.worksheets:
            sheet_name = sheet.title
            for template_sheet_name, checker in self.sheet_name_checkers.items():
                if self.accept_only_visible_sheets and sheet.sheet_state == "hidden":
                    continue
                if checker(sheet_name):
                    name_mapping[sheet_name] = template_sheet_name
                    break
        for sheet_name, template_sheet_name in name_mapping.items():
            workbook[sheet_name].title = template_sheet_name
        return workbook

    def process_excel(self, fname: str) -> ExcelProcessingResult:
        wb = openpyxl.load_workbook(fname, data_only=True)
        return self.process_workbook(wb)

    def __str__(self) -> str:
        processor = self.deref(None)
        return str(processor)


@dataclass
class ExcelProcessorFactory:
    locator_factory: LocatorFactory
    assumption_factory: AssumptionFactory
    parser_factory: ParserFactory
    validator_factory: ValidatorFactory
    table_end_condition_factory: TableEndConditionFactory

    @classmethod
    def default(cls,
                extra_locators: Optional[Dict[str, Type[Locator]]] = None,
                extra_assumptions: Optional[Dict[str, Type[Assumption]]] = None,
                extra_parsers: Optional[Dict[str, Type[Parser]]] = None,
                extra_validators: Optional[Dict[str, Type[Validator]]] = None,
                extra_table_end_conditions: Optional[Dict[str, Type[TableEndCondition]]] = None
                ):
        return ExcelProcessorFactory(
            locator_factory=LocatorFactory.default(extras=extra_locators),
            assumption_factory=AssumptionFactory.default(extras=extra_assumptions),
            parser_factory=ParserFactory.default(extras=extra_parsers),
            validator_factory=ValidatorFactory.default(extras=extra_validators),
            table_end_condition_factory=TableEndConditionFactory.default(extras=extra_table_end_conditions)
        )

    def create_extraction_task(
            self, spec: CellExtractionSpec) -> CellExtractionTask:
        try:
            return CellExtractionTask(
                key=spec.key,
                locator=self.locator_factory.create_from_spec(
                    spec=spec.locator),
                assumptions={
                    k: self.assumption_factory.create_from_spec(sp) for k, sp in spec.assumptions.items()},
                parser=self.parser_factory.create_from_spec(
                    spec=spec.parser),
                validators={
                    k: self.validator_factory.create_from_spec(sp) for k, sp in spec.validations.items()},
                fallback=spec.fallback,
                metadata=spec.apv.metadata)
        except ExcoException as e:
            raise ExtractionTaskCreationException(
                f'Unable to create ExtractionTask for {spec.key} cf {spec.source.describe()}') from e

    def create_table_extraction_task(
            self, spec: TableExtractionSpec) -> TableExtractionTask:
        columns = {k: self.create_extraction_task(CellExtractionSpec(v))
                   for k, v in spec.columns.items()}
        try:
            return TableExtractionTask(
                key=spec.key,
                locator=self.locator_factory.create_from_spec(
                    spec=spec.locator),
                columns=columns,
                end_condition=EndConditionCollection.from_spec(
                    spec.end_conditions,
                    factory=self.table_end_condition_factory),
                item_direction=spec.item_direction)
        except ExcoException as e:
            raise TableExtractionTaskCreationException(
                f'Unable to create TableExtractionTask for {spec.key} cf\n {spec.source.describe()}') from e

    def create_derefed_processor_from_spec(self,
                                           spec: ExcelProcessorSpec) -> ExcelDerefedProcessor:
        cell_tasks = {}
        for cl, specs in spec.cell_specs.items():
            cell_tasks[cl] = [
                self.create_extraction_task(spec) for spec in specs]

        table_tasks = {}
        for cl, specs in spec.table_specs.items():
            table_tasks[cl] = [
                self.create_table_extraction_task(spec) for spec in specs]

        return ExcelDerefedProcessor(cell_processors=cell_tasks,
                                     table_processors=table_tasks)

    def create_from_spec(self, spec: ExcelProcessorSpec,
                         sheet_name_checkers: Optional[SheetNameAliasCheckers] = None,
                         accept_only_visible_sheets: bool = False) -> ExcelProcessor:
        return ExcelProcessor(spec=spec, factory=self,
                              sheet_name_checkers=sheet_name_checkers,
                              accept_only_visible_sheets=accept_only_visible_sheets)

    def create_from_template_excel(self,
                                   fname: str,
                                   sheet_name_checkers: Optional[SheetNameAliasCheckers] = None,
                                   accept_only_visible_sheets: bool = False
                                   ) -> ExcelProcessor:
        workbook = openpyxl.load_workbook(fname, data_only=True)
        return self.create_from_template_workbook(workbook,
                                                  sheet_name_checkers=sheet_name_checkers,
                                                  accept_only_visible_sheets=accept_only_visible_sheets)

    def create_from_template_workbook(
            self,
            workbook: Workbook,
            sheet_name_checkers: Optional[SheetNameAliasCheckers] = None,
            accept_only_visible_sheets: bool = False
    ) -> ExcelProcessor:
        spec = ExcelProcessorSpec.from_workbook_template(workbook)
        return self.create_from_spec(spec,
                                     sheet_name_checkers=sheet_name_checkers,
                                     accept_only_visible_sheets=accept_only_visible_sheets)
